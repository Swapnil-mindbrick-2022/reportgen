from flask import Flask,render_template,request,send_file;
from flask_wtf import FlaskForm;
from wtforms import FileField,SubmitField;
import pandas as pd
import logging as logger
from werkzeug.utils import secure_filename
from openpyxl.styles import Border, Side, Alignment,PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import re
import os
import io




app = Flask(__name__)
app.config['SECRET_KEY'] = 'supersecretkey'
app.config['UPLOAD_FOLDER'] = 'static/files'



class UploadFileForm(FlaskForm):
    file = FileField("File")
    submit = SubmitField("Upload File")


@app.route('/',methods = ['GET','POST'])
@app.route('/download')


def index():
        form = UploadFileForm()
      



# @app.route('/upload', methods=['GET', 'POST'])


        
        if form.validate_on_submit():
            

            file = form.file.data
            df = pd.read_excel(file)
            unique_surveyor_names = df["Surveyor Name"].unique()

            # Create a dictionary to store the results for each surveyor name
            results = {}

            # Loop through each unique surveyor name
            for surveyor_name in unique_surveyor_names:
                # Filter the dataframe to only include data for the current surveyor name
                surveyor_data = df[df["Surveyor Name"] == surveyor_name]

                # Group the data by surveyor name and calculate the number of samples
                samples = len(surveyor_data)

                # Group the data by surveyor name and calculate the total audio duration
                duration = surveyor_data["Audio Duration (in secs)"].sum()

                # Calculate the starting time and ending time for each surveyor
                start_time = surveyor_data["Timestamp"].min()
                end_time = surveyor_data["Timestamp"].max()

                # Group the data by gender and calculate the percentage of male and female
                gender_group = surveyor_data.groupby("Gender").count()["Timestamp"]
                #print(gender_group)
                gender_percentage = (gender_group / len(surveyor_data) * 100).astype(float)
                
                def count_valid_contact_numbers(x):
                    pattern = re.compile(r"^\d{10}$")
                    mask = x.apply(lambda x: not bool(pattern.match(str(x))))
                    count = mask.sum()
                    return count
                
                invalid_numbers = count_valid_contact_numbers(surveyor_data['Contact Number'])

                
                
                duplicate = surveyor_data.duplicated('Location').sum()

                # Group the data by age group and calculate the percentage of each age group
                age_group = surveyor_data.groupby("Age Group").count()["Timestamp"].astype(float)
                age_percentage = (age_group / len(surveyor_data) * 100)
                print(age_group)

                # Group the data by political party and calculate the percentage of each party
                party_group = surveyor_data.groupby('ఇప్పటికి ఇప్పుడు ఎన్నికలు  జరిగితే మీరు ఏ పార్టీ కి మద్దతు ఇస్తారు  ?').count()["Timestamp"]
                party_percentage = (party_group / len(surveyor_data) * 100).astype(float)
                
                party_group1 = surveyor_data.groupby('మీ MLA పరిపాలన పట్ల మీ అభిప్రాయం?').count()["Timestamp"]
                
                party_percentage2 = (party_group1 / len(surveyor_data) * 100).astype(float)
                party_group2 = surveyor_data.groupby('వైయెస్ జగన్మోహన్ రెడ్డిగారి పరిపాలన పట్ల మీ అభిప్రాయం ఏమిటి?').count()["Timestamp"]
                party_percentage3 = (party_group2 / len(surveyor_data) * 100).astype(float)

                # Save the results in the dictionary
                results[surveyor_name] = {
                    "NO OF SAMPLES": samples,
                    "DURATION": duration,
                    "STARTING TIME": start_time,
                    "ENDING TIME": end_time,
                    "FEMALE": gender_percentage.get("Female", 0),
                    "MALE": gender_percentage.get("Male", 0),
                    "DUPLICATE LOCATION":duplicate,
                    'INVALID CONTACT': invalid_numbers,
                    "18-30": age_percentage.get("18-30", 0),
                    "30-45": age_percentage.get("30-45", 0),
                    "45-60": age_percentage.get("45-60", 0),
                    "60+": age_percentage.get("60+", 0),
                    "YSRCP": party_percentage.get("YSRCP", 0),
                    "TDP": party_percentage.get("TDP",0),
                    "JSP": party_percentage.get("JSP", 0),
                    "BJP": party_percentage.get("BJP", 0),
                    "INC": party_percentage.get("INC", 0),
                    "Not Decided": party_percentage.get("Not Decided", 0),
                    "బాగుంది.":party_percentage3.get('బాగుంది' ,0),
                    "బాగోలేదు.":party_percentage3.get('బాగోలేదు' ,0),
                    "బాగా చేస్తున్నారు" : party_percentage2.get("బాగా చేస్తున్నారు" , 0), 
                    "బాగా చేయడం లేదు":  party_percentage2.get("బాగా చేయడం లేదు"  , 0),
                    
                    }
                
                #results[surveyor_name]['INVALID CONTACT'] = invalidnum.Contact_Number.apply(count_valid_contact_numbers).sum()
                    
                def color_format(val):
                    text_color = 'black'
                    try:
                        
                        if val >= "60":
                            color = 'red'
                            font_weight = 'bold'
                            text_color = 'white'
                        else:
                            color = 'white'
                            font_weight = 'normal'
                        return 'background-color: %s; font-weight: %s; color: %s' % (color, font_weight,text_color)
                    except ValueError:
                        return ''
                    
                def color_format2(val):
                    text_color = 'black'
                    try:
                        if val.endswith("%") and float(val.strip("%")) >= 30:
                            color = 'red'
                            font_weight = 'bold'
                            text_color = 'white'
                        else:
                            color = 'white'
                            font_weight = 'normal'
                        return 'background-color: %s; font-weight: %s; color: %s' % (color, font_weight,text_color)
                    except ValueError:
                        return ''
                def color_format3(val):
                    text_color = 'black'
                    try:
                        if val.endswith("%") and float(val.strip("%")) >= 50:
                            color = 'red'
                            font_weight = 'bold'
                            text_color = 'white'
                        else:
                            color = 'white'
                            font_weight = 'normal'
                        return 'background-color: %s; font-weight: %s; color: %s' % (color, font_weight,text_color)
                    except ValueError:
                        return ''
                def color_format4(val):
                    text_color = 'black'
                    try:
                        if val.endswith("%") and float(val.strip("%")) >= 60:
                            color = 'red'
                            font_weight = 'bold'
                            text_color = 'white'
                        else:
                            color = 'white'
                            font_weight = 'normal'
                        return 'background-color: %s; font-weight: %s; color: %s' % (color, font_weight,text_color)
                    except ValueError:
                        return ''
            
                def duration_format(val):
                    text_color = 'black'
                    if val >= '05:00:00' :
                        color = 'red'
                        font_weight = 'bold'
                        text_color = 'white'
                    else:
                        color= 'white'
                        font_weight = 'normal'
                    
                    return 'background-color: %s; font-weight: %s; color: %s' % (color, font_weight,text_color)
                
                def duplicate_location(val):
                    text_color = 'black'
                    try:
                        
                        if val >= 5:
                            color = 'red'
                            font_weight = 'bold'
                            text_color = 'white'
                            
                        else:
                            color = 'white'
                            font_weight = 'normal'
                        return 'background-color: %s; font-weight: %s; color: %s' % (color, font_weight,text_color)
                    except ValueError:
                        return ''
                def invalid_number(val):
                    text_color = 'black'
                    try:
                        
                        if val >= 5:
                            color = 'red'
                            font_weight = 'bold'
                            text_color = 'white'
                        else:
                            color = 'white'
                            font_weight = 'normal'
                        return 'background-color: %s; font-weight: %s; color: %s' % (color, font_weight,text_color)
                    except ValueError:
                
                        return ''
                
                results_df = pd.DataFrame.from_dict(results, orient='index')
                
                #results_df = results_df.transpose()
                results_df.reset_index(inplace=True)
                
                results_df.rename(columns={"index": "Surveyor Name"}, inplace=True)
                results_df['BJP'] = results_df['BJP'].apply(lambda x: "{:.0f}%".format(x))
                results_df['INC'] = results_df['INC'].apply(lambda x: "{:.0f}%".format(x))
                results_df['JSP'] = results_df['JSP'].apply(lambda x: "{:.0f}%".format(x))
                results_df['TDP'] = results_df['TDP'].apply(lambda x: "{:.0f}%".format(x))
                results_df['YSRCP'] = results_df['YSRCP'].apply(lambda x: "{:.0f}%".format(x))
                results_df['Not Decided'] = results_df['Not Decided'].apply(lambda x: "{:.0f}%".format(x))

                results_df['18-30']= results_df['18-30'].apply(lambda x: "{:.0f}%".format(x))
                results_df['30-45']= results_df['30-45'].apply(lambda x: "{:.0f}%".format(x))
                results_df['45-60']= results_df['45-60'].apply(lambda x: "{:.0f}%".format(x))
                results_df['60+']= results_df['60+'].apply(lambda x: "{:.0f}%".format(x))
                results_df['MALE']= results_df['MALE'].apply(lambda x: "{:.0f}%".format(x))
                results_df['FEMALE']= results_df['FEMALE'].apply(lambda x: "{:.0f}%".format(x))
                results_df['బాగా చేయడం లేదు']= results_df['బాగా చేయడం లేదు'].apply(lambda x: "{:.0f}%".format(x))
                results_df['బాగా చేస్తున్నారు']= results_df['బాగా చేస్తున్నారు'].apply(lambda x: "{:.0f}%".format(x))
                results_df['బాగోలేదు.']= results_df['బాగోలేదు.'].apply(lambda x: "{:.0f}%".format(x))
                results_df['బాగుంది.']= results_df['బాగుంది.'].apply(lambda x: "{:.0f}%".format(x))


                results_df['STARTING TIME'] = results_df['STARTING TIME'].apply(lambda x: datetime.strptime(x, '%Y-%m-%d %H:%M:%S.%f'))
                results_df['ENDING TIME'] = results_df['ENDING TIME'].apply(lambda x: datetime.strptime(x, '%Y-%m-%d %H:%M:%S.%f'))
                
                results_df['DURATION'] = results_df['ENDING TIME'] - results_df['STARTING TIME']
                
                results_df['DURATION'] = results_df['DURATION'].apply(lambda x: f"{x.days * 24 + x.seconds // 3600:0>2}:{(x.seconds % 3600) // 60:0>2}:{x.seconds % 60:0>2}")


                    
                
                results_df = results_df.style.applymap(color_format, subset=['MALE', 'FEMALE']) \
                    .applymap(duration_format, subset=['DURATION']) \
                    .applymap(duplicate_location, subset=['DUPLICATE LOCATION']) \
                    .applymap(invalid_number, subset=['INVALID CONTACT']) \
                    .applymap(color_format2, subset=['18-30','30-45','45-60','60+'])\
                    .applymap(color_format3, subset=['YSRCP','TDP','JSP','BJP','INC','Not Decided'])\
                    .applymap(color_format4, subset=['బాగుంది.','బాగోలేదు.' ])\
                    .applymap(color_format4, subset=['బాగా చేస్తున్నారు', 'బాగా చేయడం లేదు' ])
                    
                

                #results_df.to_excel(r"C:\Users\dell\Downloads\Sai_Swapnill Cons_Results.xlsx")

                # Save the DataFrame to an Excel file
                writer = pd.ExcelWriter("result.xlsx", engine='openpyxl')
                results_df.to_excel(writer, index=False)
                
                
                
                # Get the active worksheet
                worksheet = writer.book.active
                worksheet.freeze_panes = worksheet.cell(1,2)
            


        
                
                # Set the column width to automatically adjust to the size of the contents in the column
                for column_cells in worksheet.columns:
                    length = max(len(str(cell)) for cell in column_cells)
                    worksheet.column_dimensions[column_cells[0].column_letter].width = length
                worksheet.insert_rows(1, 1)
                worksheet['C1']="TIME"
                worksheet.merge_cells('C1:E1') 
                worksheet['F1']="GENDER"
                worksheet.merge_cells('F1:G1') 
                
                worksheet['J1']="AGE GROUP"
                worksheet.merge_cells('J1:M1') 
                
                worksheet['N1'] = 'ఇప్పటికి ఇప్పుడు ఎన్నికలు  జరిగితే మీరు ఏ పార్టీ కి మద్దతు ఇస్తారు  ?'
                worksheet.merge_cells('N1:S1') 
                worksheet['T1']= "వైయెస్ జగన్మోహన్ రెడ్డిగారి పరిపాలన పట్ల మీ అభిప్రాయం ఏమిటి?""మీ MLA పరిపాలన పట్ల మీ అభిప్రాయం?"
                worksheet.merge_cells('T1:U1') 
                worksheet['V1'] = "మీ MLA పరిపాలన పట్ల మీ అభిప్రాయం?"
                
                worksheet.merge_cells('V1:W1')
                fill_colors = ["D8E4BCFF", "D8E4BCFF", "D8E4BCFF", "D8E4BCFF", "D8E4BCFF","D8E4BCFF",]
                for i, merged_cell_range in enumerate(worksheet.merged_cell_ranges):
                    range_string = str(merged_cell_range)
                    merged_row = worksheet[range_string]
                    for row in merged_row:
                        for cell in row:
                            cell.fill = PatternFill(start_color=fill_colors[i], end_color=fill_colors[i], fill_type="solid")
                        

            

                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = cell.border + Border(left=Side(style='thin'), 
                                                        right=Side(style='thin'), 
                                                        top=Side(style='thin'), 
                                                        bottom=Side(style='thin'))
                        cell.alignment = Alignment(horizontal='center')
                
                # Set the background color of the first row (the column names)
                for cell in worksheet[2]:
                    cell.fill = PatternFill(start_color="B8CCE4FF", end_color="B8CCE4FF", fill_type = "solid")
                #Add filter to each column
                worksheet.auto_filter.ref = "A2:%s2" % (get_column_letter(worksheet.max_column))
            # data = Reference(worksheet, min_col=2, min_row=1, max_row=worksheet.max_row, max_col=2)


                # Save the changes to the Excel file
                writer.save()

            print("Analysis complete!")

            # Convert categorical columns to numeric
            
            
            # Return the file for download
            
            return send_file("result.xlsx",
                            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            as_attachment=True)

           
            # rendered=render_template('data.html',res = res,length = length)
            # df1 = pd.DataFrame({'Data': [rendered]})
            # output = io.BytesIO()
            # writer = pd.ExcelWriter(output, engine='xlsxwriter')
            # df1.to_excel(writer, index=False, sheet_name='Sheet1')
            # writer.save()
            # output.seek(0)
            # return send_file(output,
            #                 attachment_filename='combined.xlsx',
            #                 as_attachment=True)
            
            # result.to_excel("swapnil_New.xlsx", index=False)
            #return render_template('data.html',res = res,length = length)
            #result.to_excel(os.path.join(app.config['UPLOAD_FOLDER'], "swapnil_New.xlsx"), index=False)
            
            # return "Report Generated Successfully"

        return render_template('index.html',form = form)
            



if __name__ == '__main__':
    app.run()